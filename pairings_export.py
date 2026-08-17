"""
Golf Pool - Pairings Export
----------------------------
Reads the Master workbook's Pairings sheet (built from Intake, verified against
Master Column P) and produces pairings.json for the live-scoring pipeline.

Usage:
    python pairings_export.py Master_Spreadsheet_FedEx_Cup_13.xlsx pairings.json

Output format (list of dicts, one per Pairing):
    {
        "entrant_full_name": "Amy Foley",
        "label": "Amy F",
        "pairing_id": "Amy F-1",
        "golfers": ["Scottie Scheffler", "Chris Gotterup", "Viktor Hovland"],
        "side_pool": false
    }

entrant_full_name and label are both included per the leaderboard's search
requirement (entrants may search by either). Full names are pulled from the
Master sheet's compact Entrant First/Last columns (I, J), matched to the
Pairings sheet by label in column K. (Corrected 2026-08-14 — columns had
shifted one over from the H/I/J originally assumed here; see Master row 1
headers if this ever needs reconfirming.)

side_pool reflects Master column AH ("Side Pool") on a PER-PAIRING basis:
each pairing's own AH cell is checked individually — a "Yes" on that
specific row puts that specific pairing in the Side Pool, and nothing else
is implied about that entrant's other pairings. There is no cap on how many
pairings per entrant can be marked; Dale controls that per tournament simply
by how many rows he marks "Yes" for a given entrant, with no code change
required. (Rewritten 2026-08-17 — the prior version checked only whether an
entrant's label appeared ANYWHERE with "Yes" and then flagged ALL of that
entrant's pairings as side_pool, ignoring which specific rows were actually
marked. That was an entrant-level, all-or-nothing behavior and did not
support a per-pairing selection at all — see Reference Document Section 2.7
for how this was discovered and why it was changed.)

Blocks with fewer than 3 golfers (e.g. an entrant whose data is still pending)
are skipped and reported, not written as partial/broken entries.
"""

import sys
import json
import openpyxl


def build_label_to_name_map(master_ws):
    """Forward-fill Master!I (First), J (Last) against K (label) -> {label: 'First Last'}"""
    mapping = {}
    cur_first = cur_last = None
    r = 2
    while master_ws.cell(r, 11).value:  # column K = label
        i = master_ws.cell(r, 9).value
        j = master_ws.cell(r, 10).value
        k = master_ws.cell(r, 11).value
        if i:
            cur_first, cur_last = i.strip(), j.strip()
        if k not in mapping:
            mapping[k] = f"{cur_first.strip()} {cur_last.strip()}"
        r += 1
    return mapping


def build_side_pool_pairing_set(master_ws):
    """Scan Master!K (label) + L (Pairing #) against AH (Side Pool), row by row,
    -> set of (label, pairing_num) tuples whose OWN row is marked 'Yes'.

    This is deliberately per-row, not per-entrant: an entrant can have any
    number of their pairings marked (0 up to all of them), and only the
    specific pairings marked 'Yes' are included. Dale adjusts how many
    pairings per entrant qualify for the Side Pool each tournament purely by
    how many rows he marks — no code or cap here needs to change to support
    that.
    """
    marked = set()
    r = 2
    while master_ws.cell(r, 11).value:
        k = master_ws.cell(r, 11).value
        l = master_ws.cell(r, 12).value
        ah = master_ws.cell(r, 34).value  # column AH = Side Pool
        if isinstance(ah, str) and ah.strip().lower() == 'yes':
            marked.add((k, l))
        r += 1
    return marked


def export_pairings(workbook_path, output_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    master_ws = wb['Master']
    pairings_ws = wb['Pairings']

    label_to_name = build_label_to_name_map(master_ws)
    side_pool_pairings = build_side_pool_pairing_set(master_ws)

    records = []
    skipped = []
    r = 3  # Pairings sheet data starts at row 3 (row1=title, row2=headers)
    max_row = pairings_ws.max_row
    while r <= max_row:
        label = pairings_ws.cell(r, 1).value
        pairing_num = pairings_ws.cell(r, 2).value
        if label:
            g1 = pairings_ws.cell(r, 4).value
            g2 = pairings_ws.cell(r + 1, 4).value
            g3 = pairings_ws.cell(r + 2, 4).value
            golfers = [g for g in (g1, g2, g3) if g]
            if len(golfers) != 3:
                skipped.append((label, pairing_num))
            else:
                full_name = label_to_name.get(label, label)
                records.append({
                    "entrant_full_name": full_name,
                    "label": label,
                    "pairing_id": f"{label}-{pairing_num}",
                    "golfers": golfers,
                    "side_pool": (label, pairing_num) in side_pool_pairings,
                })
            r += 4  # advance to next block (3 golfer rows + 1 blank separator)
        else:
            r += 1

    with open(output_path, 'w') as f:
        json.dump(records, f, indent=2)

    print(f"Wrote {len(records)} pairings to {output_path}")
    side_count = sum(1 for rec in records if rec['side_pool'])
    print(f"Side Pool pairings: {side_count}")
    if skipped:
        print(f"Skipped {len(skipped)} incomplete pairing(s) (no golfer data yet):")
        for label, pnum in skipped:
            print(f"  - {label}, Pairing {pnum}")

    return records, skipped


if __name__ == '__main__':
    workbook_path = sys.argv[1] if len(sys.argv) > 1 else 'Master_Spreadsheet_FedEx_Cup_13.xlsx'
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'pairings.json'
    export_pairings(workbook_path, output_path)
